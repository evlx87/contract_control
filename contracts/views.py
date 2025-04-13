import io
import logging

from django.contrib import messages
from django.db import IntegrityError
from django.db.models import Sum
from django.http import HttpResponse, Http404
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, TemplateView, DetailView
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook

from contracts.forms import ContractForm, PaymentDocumentForm, PaymentOrderForm, AdditionalAgreementForm
from contracts.models import Contract, AdditionalAgreement
from lib_ccportal.models import PurchaseType

logger = logging.getLogger(__name__)


# Create your views here.
class IndexView(TemplateView):
    template_name = 'contracts/index.html'
    login_url = 'registration:login'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Главная'

        if self.request.user.is_authenticated:
            context['username'] = self.request.user.username

        return context


class AddContractView(View):
    def get(self, request):
        form = ContractForm()
        logger.info("Страница добавления контракта была загружена")
        return render(request, 'contracts/purchase_add.html', {'form': form})

    def post(self, request):
        form = ContractForm(request.POST, request.FILES)
        logger.debug("Отправленные данные формы: %s", request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Контракт успешно добавлен')
            return redirect('/contracts/')
        else:
            logger.error("Форма контракта не валидна: %s", form.errors)
            messages.error(
                request,
                'Ошибка при добавлении контракта. Проверьте введенные данные')
            return render(request,
                          'contracts/purchase_add.html',
                          {'form': form})


class PurchaseListView(ListView):
    model = Contract
    template_name = 'contracts/purchases_list.html'
    context_object_name = 'purchases'

    def get_queryset(self):
        queryset = super().get_queryset()
        year = self.request.GET.get('year')
        subject = self.request.GET.get('subject')

        if year:
            queryset = queryset.filter(service_end_date__year=year)
        else:
            year = timezone.now().year
            queryset = queryset.filter(service_end_date__year=year)
        if subject:
            queryset = queryset.filter(contract_subject=subject)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_year = timezone.now().year
        years = Contract.objects.dates('service_end_date', 'year')
        context['years'] = years
        context['selected_year'] = self.request.GET.get('year', str(current_year))
        subjects = Contract.objects.values_list('contract_subject', flat=True).distinct()
        context['subjects'] = subjects
        context['selected_subject'] = self.request.GET.get('subject', '')

        # Получаем все объекты после фильтрации
        purchases = self.get_queryset()

        # Вычисляем сумму contract_amount через aggregate
        total_contract_amount = purchases.aggregate(Sum('contract_amount'))['contract_amount__sum'] or 0

        # Вычисляем сумму total_pp_issued_amount вручную
        total_payment_amount = sum(purchase.total_pp_issued_amount for purchase in purchases) or 0

        context['total_contract_amount'] = total_contract_amount
        context['total_payment_amount'] = total_payment_amount

        return context


class ContractDetailView(DetailView):
    model = Contract
    template_name = 'contracts/contract_detail.html'
    context_object_name = 'contract'

    def get_object(self, queryset=None):
        obj = super().get_object(queryset)
        if not obj:
            raise Http404("Контракт не найден.")
        return obj

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['additional_agreements'] = self.object.additional_agreements.all()
        # Вычисляем остатки
        contract = self.object
        context['remaining_after_payments'] = contract.contract_amount - contract.total_issued_amount
        context['remaining_after_contract'] = contract.contract_amount - contract.total_pp_issued_amount
        return context


class ContractEditView(View):
    template_name = 'contracts/contract_edit.html'

    def get_object(self, contract_id):
        return get_object_or_404(Contract, pk=contract_id)

    def get(self, request, contract_id):
        contract = self.get_object(contract_id)
        form = ContractForm(instance=contract)
        additional_agreements = contract.additional_agreements.all()
        return render(request, self.template_name, {
            'form': form,
            'contract': contract,
            'additional_agreements': additional_agreements
        })

    def post(self, request, contract_id):
        contract = self.get_object(contract_id)
        form = ContractForm(request.POST, request.FILES, instance=contract)

        if form.is_valid():
            if 'contract_file' in request.FILES:
                contract.contract_file = request.FILES['contract_file']
            form.save()
            logger.info(f"Контракт с ID {contract.id} успешно отредактирован")
            return redirect('contract-detail', pk=contract.id)

        return render(request, self.template_name, {
            'form': form,
            'contract': contract,
            'additional_agreements': contract.additional_agreements.all()
        })


class ContractDeleteView(View):
    template_name = 'contracts/contract_delete.html'

    def get_object(self, pk):
        return get_object_or_404(Contract, pk=pk)

    def get(self, request, pk):
        contract = self.get_object(pk)
        return render(request, self.template_name, {'contract': contract})

    def post(self, request, pk):
        contract = self.get_object(pk)
        contract.delete()
        messages.success(request, 'Данные о закупке были успешно удалены.')
        logger.info(f"Контракт с ID {pk} был удален")
        return redirect('contracts:purchase_list')


class AddPaymentDocView(View):
    template_name = 'contracts/payment_doc_add.html'

    def get_contract(self, contract_id):
        return get_object_or_404(Contract, id=contract_id)

    def get(self, request, contract_id=None):
        contract = None
        if contract_id:
            contract = self.get_contract(contract_id)
            form = PaymentDocumentForm(initial={'contract': contract})
        else:
            form = PaymentDocumentForm()
        logger.info('Запрошена страница для добавления документа оплаты')
        return render(
            request, self.template_name, {
                'form': form, 'contract': contract})

    def post(self, request, contract_id):
        contract = self.get_contract(contract_id)
        form = PaymentDocumentForm(request.POST, request.FILES)

        if form.is_valid():
            payment_document = form.save(commit=False)
            payment_document.contract = contract
            try:
                payment_document.save()
                logger.info(
                    f"Создан новый документ оплаты для контракта с ID {contract_id}")
                return redirect('contracts:contract-detail', pk=contract.id)
            except IntegrityError as e:
                logger.error(f"Ошибка при сохранении документа: {str(e)}")
                # Добавить общее сообщение об ошибке
                form.add_error(
                    None,
                    'Произошла ошибка при сохранении документа. Пожалуйста, попробуйте снова.')
        else:
            logger.warning(
                'Форма документа оплаты содержит ошибки: %s',
                form.errors)

        # Если форма невалидна, отображаем ее снова с ошибками
        return render(
            request, self.template_name, {
                'form': form, 'contract': contract})


class AddPaymentOrderView(View):
    template_name = 'contracts/payment_order_add.html'

    def get(self, request, contract_id=None):
        contract = get_object_or_404(Contract, id=contract_id)
        form = PaymentOrderForm(initial={'contract': contract})
        logger.info(
            f"Отображена страница для добавления платежного поручения для контракта с ID {contract_id}")
        return render(
            request, self.template_name, {
                'form': form, 'contract': contract})

    def post(self, request, contract_id):
        contract = get_object_or_404(Contract, id=contract_id)
        form = PaymentOrderForm(request.POST, request.FILES)
        if form.is_valid():
            payment_order = form.save(commit=False)
            payment_order.contract = contract
            try:
                payment_order.save()
                logger.info(
                    f"Создан новый документ платежного поручения для контракта с ID {contract_id}")
                return redirect('contracts:contract-detail', pk=contract.id)
            except IntegrityError as e:
                logger.error(f"Ошибка при сохранении документа: {str(e)}")
                # Добавить общее сообщение об ошибке
                form.add_error(
                    None,
                    'Произошла ошибка при сохранении документа. Пожалуйста, попробуйте снова.')
            logger.info(
                f"Создано новое платежное поручение для контракта с ID {contract_id}")
            return redirect(
                'contracts:contract-detail',
                pk=contract.id)
        else:
            logger.warning('Форма платежного поручения содержит ошибки')

        return render(
            request, self.template_name, {
                'form': form, 'contract': contract})


class JournalListView(ListView):
    model = Contract
    template_name = 'contracts/journal_list.html'
    context_object_name = 'purchases'

    def get_queryset(self):
        queryset = super().get_queryset()
        year = self.request.GET.get('year')
        purchase_type = self.request.GET.get('purchase_type')

        if year:
            queryset = queryset.filter(contract_date__year=year)
        if purchase_type:
            queryset = queryset.filter(purchase_type__code=purchase_type)

        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['page_title'] = 'Журнал регистрации контрактов'
        # Получаем список годов
        years = Contract.objects.dates('contract_date', 'year', order='DESC')
        context['years'] = years
        # Получаем список типов закупок
        purchase_types = PurchaseType.objects.values_list('code', flat=True).distinct()
        context['purchase_types'] = purchase_types
        # Передаем выбранные значения фильтров
        context['selected_year'] = self.request.GET.get('year', '')
        context['selected_purchase_type'] = self.request.GET.get('purchase_type', '')
        return context


# Представление для экспорта в Excel
class ExportToExcelView(View):

    def get(self, request):
        # Получение всех объектов модели контракта
        purchases = Contract.objects.all().order_by('-id')

        # Данные для заполнения таблицы
        data = [['Дата контракта',
                 'Номер контракта',
                 'Контрагент',
                 'Предмет контракта',
                 'Срок действия контракта',
                 'Дата начала услуг/поставки',
                 'Дата окончания услуг/поставки',
                 'Сумма контракта',
                 'Тип закупки']]

        for purchase in purchases:
            data.append([
                purchase.contract_date.strftime('%d.%m.%Y'),
                purchase.contract_number,
                purchase.supplier,
                purchase.contract_subject,
                purchase.contract_duration.strftime('%d.%m.%Y'),
                purchase.service_start_date.strftime('%d.%m.%Y'),
                purchase.service_end_date.strftime('%d.%m.%Y'),
                purchase.contract_amount,
                str(purchase.purchase_type),
            ])

        # Создание новой рабочей книги
        wb = Workbook()
        ws = wb.active

        # Заполнение данными
        for row in data:
            ws.append(row)

        # Установка ширины колонок
        dim_holder = {}
        for col in range(len(data[0])):
            for row in range(1, len(data) + 1):
                column_letter = get_column_letter(col + 1)
                cell_value = str(ws.cell(row=row, column=col + 1).value)
                try:
                    dim_holder[column_letter].append(len(cell_value))
                except KeyError:
                    dim_holder[column_letter] = [len(cell_value)]

        for col, widths in dim_holder.items():
            max_width = max(widths)
            ws.column_dimensions[col].width = max_width + 3

        # Сохранение файла в память
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Отправка файла пользователю
        response = HttpResponse(
            content=output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="journal.xlsx"'
        return response


class AddAdditionalAgreementView(View):
    def get(self, request, contract_id):
        contract = get_object_or_404(Contract, id=contract_id)
        form = AdditionalAgreementForm()
        return render(request, 'contracts/additional_agreement_add.html', {'form': form, 'contract': contract})

    def post(self, request, contract_id):
        contract = get_object_or_404(Contract, id=contract_id)
        form = AdditionalAgreementForm(request.POST, request.FILES)
        if form.is_valid():
            additional_agreement = form.save(commit=False)
            additional_agreement.contract = contract
            additional_agreement.save()
            messages.success(request, 'Дополнительное соглашение успешно добавлено.')
            return redirect('contracts:contract-detail', pk=contract.id)
        else:
            logger.error("Ошибка при добавлении дополнительного соглашения: %s", form.errors)
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме.')
            return render(request, 'contracts/additional_agreement_add.html', {'form': form, 'contract': contract})


class AdditionalAgreementDetailView(DetailView):
    model = AdditionalAgreement
    template_name = 'contracts/additional_agreement_detail.html'
    context_object_name = 'agreement'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['contract'] = self.object.contract  # Добавляем контракт в контекст
        return context